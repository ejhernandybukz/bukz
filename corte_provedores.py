import streamlit as st
import pandas as pd
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from openpyxl.utils import get_column_letter
from email.mime.text import MIMEText
from io import BytesIO
import zipfile



def corte_provedores():
    st.title("Corte Proveedores")
    
    st.markdown("<h3>Archivo con correo de proveedores</h3>", unsafe_allow_html=True)
    proveedores_df_1 = st.file_uploader("El archivo debe tener las columnas: Proveedores y Correo Medellin", type=["xlsx"], key="archivo_proveedores")
    
    
    if 'envio_proveedores' not in st.session_state or st.session_state['envio_proveedores'].empty:
        st.session_state['envio_proveedores'] = pd.DataFrame(columns=['Proveedor', 'Estado'])
    
    st.markdown("<h3>Archivo con ventas mensuales</h3>", unsafe_allow_html=True)
    uploaded_file_ventas_mesuales = st.file_uploader("El archivo debe tener las columnas: product_title, variant_sku, product_vendor, pos_location_name, net_quantity", type=["xlsx"], key="archivo_productos")
    
    if uploaded_file_ventas_mesuales is not None and proveedores_df_1 is not None:
        if st.button("Continuar"):
            st.session_state['archivo_cargado'] = True
    
    if st.session_state['archivo_cargado']:
        proveedores_df = pd.read_excel(proveedores_df_1)
        
        grouped_dfs = None  # Inicializa la variable aquí
    
        # Widgets para ingresar datos del usuario
        st.session_state.smtp_user = st.text_input("Ingrese el usuario SMTP:", value=st.session_state.smtp_user)
        st.session_state.smtp_password = st.text_input("Ingrese la contraseña SMTP:", type="password", value=st.session_state.smtp_password)
    
        mes = st.text_input("Ingrese el mes:")  # No hay valor predeterminado
        año = st.text_input("Ingrese el año:")  # No hay valor predeterminado
        
        if 'nombre_remitente' not in st.session_state:
            st.session_state['nombre_remitente'] = "Sebastian Barrios - Bukz"  # Valor por defecto
    
        #Otros campos de texto
        remitente_default = "Sebastian Barrios - Bukz"
        remitente_otros = "Otro (escriba abajo)"
        nombre_remitente_seleccion = st.selectbox("Seleccione el nombre del remitente:", [remitente_default, remitente_otros])
    
        # Actualizar el valor de nombre_remitente en st.session_state
        if nombre_remitente_seleccion == remitente_otros:
            if 'nombre_remitente_personalizado' not in st.session_state:
                st.session_state['nombre_remitente_personalizado'] = ''
            st.session_state['nombre_remitente'] = st.text_input("Ingrese el nombre del remitente:", key="nombre_remitente_personalizado")
        else:
            st.session_state['nombre_remitente'] = remitente_default
            
        # Firma
        firma_default = "Sebastian Barrios - Analista de Operaciones"
        firma_otros = "Otra (escriba abajo)"
        firma_seleccion = st.selectbox("Seleccione la firma:", [firma_default, firma_otros])
        
        if st.button("Procesar archivo y enviar correos"):
            
            df = pd.read_excel(uploaded_file_ventas_mesuales)
            df['pos_location_name'] = df['pos_location_name'].replace({
                'Bukz Las Lomas': 'Medellin',
                'Bukz Tesoro': 'Medellin',
                'Bukz Mattelsa': 'Medellin',
                '': 'Medellin',
                'Bukz St. Patrick': 'Bogota',
                'Vassar': 'Bogota',
                'Bukz Bogota 109': 'Bogota'
                })
            
            df['pos_location_name'] = df['pos_location_name'].fillna('Medellin')
            proveedores_a_eliminar = ["603 La Gran Via", "Alejandra Márquez Villegas", "Alejandro Salazar Yusti",
                "Andina", "Books for U", "Bukz", "Bukz España", "Bukz USA", "Bukz.co",
                "Fernando Ayerbe", "Grupo Editorial Planeta", "Juan D. Hoyos Distribuciones SAS",
                "Libros de Ruta", "Luminosa", "Melon", "Penguin RandomHouse",
                "Pergamino Café", "Postobon", "Tea market", "Torrealta", "Urban",
                "Álvaro González Alorda"]
    
            df = df[~df['product_vendor'].isin(proveedores_a_eliminar)]
            
            # Función personalizada para sumar teniendo en cuenta la lógica de negativos y positivos
            def custom_sum(group):
                if all(group <= 0):  # Si todos los valores son negativos o cero, suma todos
                    return group.sum()
                else:  # Si no, suma solo los positivos y ceros
                    return group[group >= 0].sum()
                
            # Aplicar la función personalizada para la agrupación
            df_grouped = df.groupby(['product_title', 'product_vendor', 'variant_sku', 'pos_location_name'])['net_quantity'].apply(custom_sum).reset_index()
            
            def format_variant_sku(value):
                try:
                    return '{:.0f}'.format(float(value))
                except ValueError:
                    return value  # Devuelve el valor original si no se puede convertir
            
            df_grouped['variant_sku'] = df_grouped['variant_sku'].apply(format_variant_sku) 
            grouped_dfs = {vendor: sub_df for vendor, sub_df in df_grouped.groupby('product_vendor')}
            
            # Configuración del servidor de correo
            smtp_server = 'smtp.gmail.com'
            smtp_port = 587                
            # Conexión al servidor SMTP
            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            server.login(st.session_state.smtp_user, st.session_state.smtp_password)
    
            imagen_url = "https://ci3.googleusercontent.com/mail-sig/AIorK4zk7DTZK_4Nl0qLnpmzJnoAhaN3t08JpWQmDUdtbhe-nJySTGmVsdjlqZr7sVzEJzCFTSGzHY8" 
            
            # Función para ajustar el ancho de las columnas
            def ajustar_ancho_columnas(workbook, ancho):
                for sheet in workbook.sheetnames:
                    worksheet = workbook[sheet]
                    for col in worksheet.columns:
                        col_letter = get_column_letter(col[0].column)
                        worksheet.column_dimensions[col_letter].width = ancho
            
            # Función modificada para enviar un DataFrame como archivo Excel con múltiples hojas a múltiples correos
            def enviar_df_por_correo(df, emails, subject, body_message,nombre_remitente):
                # Crear un buffer en memoria para el archivo Excel
                output = BytesIO()
                
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    for ciudad in ['Medellin', 'Bogota']:
                        df_ciudad = df[df['pos_location_name'] == ciudad]
                        if not df_ciudad.empty:
                            df_ciudad.to_excel(writer, sheet_name=ciudad, index=False)
                    # No es necesario llamar a save() aquí, se maneja automáticamente.
            
                # Mover el puntero al inicio del buffer
                output.seek(0)
            
                # Crear el mensaje de correo electrónico
                msg = MIMEMultipart()
                msg['From'] = f"{nombre_remitente}"
                msg['To'] = ", ".join(emails)
                msg['Subject'] = subject
            
                # Agregar el cuerpo del mensaje
                msg.attach(MIMEText(body_message, 'html'))
            
                # Adjuntar el archivo Excel desde el buffer en memoria
                part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                part.set_payload(output.getvalue())  # Usamos getvalue() para obtener el contenido del buffer
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', 'attachment', filename=f'Corte_Ventas_{mes}.xlsx')
                msg.attach(part)
            
                # Enviar el correo
                server.send_message(msg)
            
                # Cerrar el buffer
                output.close()
                
                st.session_state['procesado'] = True
            
            # Texto del cuerpo del correo
            body_message = f"""<p>Buenas tardes,</p>
            
            <p>Espero que estés muy bien. Adjunto información del corte del mes de {mes}.</p>
            
            <p>Le recordamos que la cuenta de cobro o factura debe ser remitida <strong>exclusivamente a facturacion@bukz.co antes del día 25 del mes en curso</strong>.</p>
            
            <p> Por asuntos contables, les pedimos amablemente que envíen una factura especificando el lugar de venta, indicando si es <strong>Medellín</strong> y otra que especifique si es <strong>Bogotá</strong>, así como se especifica en los archivos del corte de venta.</p>
            
            <p>Para cualquier consulta o asunto adicional, no dude en contactar a los siguientes departamentos:</p>
            <ul>
                <li>Facturación: facturacion@bukz.co</li>
                <li>Bodega y Devoluciones: cedi@bukz.co</li>
               
            </ul>
            
            <p>Saludos cordiales,</p>
            
            <p><strong style="color: gray;">{firma_seleccion}</strong></p>
            <p><img src="{imagen_url}" alt="Logo Bukz" style="width: 150px;"></p>  
            
            """
            
            # Enviar correos a proveedores
            for proveedor, sub_df in grouped_dfs.items():
                correos_filtrados = proveedores_df[proveedores_df['Proveedores'] == proveedor]['Correo Medellin']
                if not correos_filtrados.empty:
                    correos = correos_filtrados.iloc[0].split(';')
                    try:
                        enviar_df_por_correo(sub_df, correos, f"Corte {mes} {año} - {proveedor}", body_message, st.session_state['nombre_remitente'])
                        nuevo_registro = {'Proveedor': proveedor, 'Estado': 'Enviado'}
                    except Exception as e:
                        nuevo_registro = {'Proveedor': proveedor, 'Estado': f'Fallo en el envío: {e}'}
                else:
                    nuevo_registro = {'Proveedor': proveedor, 'Estado': 'Correo no encontrado'}
            
                # Aquí usamos append y luego reasignamos el DataFrame actualizado
                nuevo_registro_df = pd.DataFrame([nuevo_registro])
                st.session_state['envio_proveedores'] = pd.concat([st.session_state['envio_proveedores'], nuevo_registro_df], ignore_index=True)
                
            
    
        if st.session_state['procesado']:
            # Mostrar el DataFrame con los resultados de los envíos
            st.dataframe(st.session_state['envio_proveedores'])
            
            # Función para convertir un DataFrame a Excel y luego codificarlo para la descarga
            def to_excel(df):
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False)
                processed_data = output.getvalue()
                return processed_data
            
            # Función para guardar un DataFrame en un buffer como archivo Excel
            def guardar_excel_en_buffer(df, nombre_archivo):
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    # Crear una hoja para cada ciudad en el archivo Excel solo si el DataFrame correspondiente no está vacío
                    for ciudad in ['Medellin', 'Bogota']:
                        df_ciudad = df[df['pos_location_name'] == ciudad]
                        if not df_ciudad.empty:
                            df_ciudad.to_excel(writer, sheet_name=ciudad, index=False)
            
                    # Verificar si se agregaron hojas al archivo
                    if not writer.sheets:
                        # Si no hay hojas, agrega una hoja con un DataFrame vacío para evitar el error
                        pd.DataFrame().to_excel(writer, sheet_name='Vacio', index=False)
            
                output.seek(0)
                return output.getvalue()
            
            # Crear un archivo ZIP en memoria
            zip_buffer = BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED, False) as zip_file:
                
                # Agregar el DataFrame de estado de envío como archivo Excel en el ZIP
                estado_envio_excel = to_excel(st.session_state['envio_proveedores'])
                zip_file.writestr('estado_envio.xlsx', estado_envio_excel)
            
                # Agregar los archivos Excel de cada proveedor en el ZIP
                if grouped_dfs is not None:
                    for proveedor, sub_df in grouped_dfs.items():
                        excel_data = guardar_excel_en_buffer(sub_df, f'{proveedor}.xlsx')
                        zip_file.writestr(f'{proveedor}.xlsx', excel_data)
            
            # Preparar el buffer para la descarga
            zip_buffer.seek(0)
            
            # Botón para descargar el archivo ZIP que incluye tanto el estado de envío como los archivos de proveedores
            st.download_button(
            label="Descargar Estado de Envíos y Archivos de Proveedores",
            data=zip_buffer,
            file_name=f"envios_y_proveedores_{mes}.zip",
            mime='application/zip')
